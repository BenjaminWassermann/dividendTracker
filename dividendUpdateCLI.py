#!/usr/bin/env python3
# dividendUpdateCLI.py

# reads a list of tickers from an excel sheet
# updates current price and annual yield
# includes advanced CLI which can generate a sheet,
# update a sheet, add a single or list of tickers,
# delete a single or list of tickers

# import statements for openpyxl yahoo_fin datetime click os and math
import openpyxl, datetime, click, os, math
from yahoo_fin import stock_info as si

# save lists of dow, nasdaq and other
dow = si.tickers_dow()
nasdaq = si.tickers_nasdaq
other = si.tickers_other

# create list of lists
exchange = [dow, nasdaq, other]

# start click group, run without command
@click.group(invoke_without_command=True)

# option for loading tickers from a different file
@click.option('--file-in', '-i', default='dividendCalc.xlsx', show_default=True)

# option for saving tickers to a different file
@click.option('--file-out', '-o', default='dividendCalc.xlsx', show_default=True)
@click.pass_context

# main option
def cli(ctx, file_in, file_out):
    """Updates all tickers in file_in and saves to file_out"""

    # get the basic path
    basicPath = os.getcwd()

    # create in and out paths
    inPath = '%s\\%s' % (basicPath, file_in)
    outPath = '%s\\%s' % (basicPath, file_out)

    # throw error if inPath doesn't exist or isn't an excel document
    if os.path.exists(inPath) and inPath[-4:]=='xlsx':

        # save filenames and absolute paths for in and out files
        ctx.obj['file_in'] = file_in
        ctx.obj['inPath'] = inPath
        ctx.obj['file_out'] = file_out
        ctx.obj['outPath'] = outPath

        # if no subcommand is called, run full sheet update
        if ctx.invoked_subcommand is None:
            click.echo('Updating tickers from %s and saving to %s...' % (file_in, file_out))
            dividendUpdate(file_in, file_out)
            click.echo('Update complete!')
            os.startfile(outPath)
        else:
            click.echo('Invoking %s...' % ctx.invoked_subcommand)
                    
    else:
        raise Exception('file-in must be an extant xlsx document')

# creates a new formatted sheet with no tickers and saves to file_out
@cli.command()
@click.pass_context
def new(ctx):
    """Creates a new formatted sheet with no tickers and saves to file_out"""
    
    file_out = ctx.obj['file_out']
    outPath = ctx.obj['outPath']
    
    click.echo('Generating new sheet and saving to %s...' % file_out)
    
    # dicitonary containing cells and values
    entries = {'A1':'Stock ID',
               'B1':'Price',
               'C1':'Yield',
               'D1':'Annual Yield',
               'E1':'$price/$annual',
               'F1':'Annual yield for $1k',
               'G1':'Updated:'}

    # create new workbook and select active sheet
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = 'dividend calc'

    # place values in cells
    for position, entry in entries.items():
        sheet[position].value = entry

    # update date
    updateDate(sheet)    

    # save sheet as toFill.xlsx and close
    book.save(file_out)
    book.close()

    # print finished statement
    click.echo('%s generated successfully' % file_out)
    os.startfile(outPath)
    

# add command adds a new ticker and updates just that ticker
@cli.command()
@click.argument('ticker')
@click.pass_context
def add(ctx, ticker):
    """Adds a new ticker to file_in and saves to file_out"""
    
    file_in = ctx.obj['file_in']
    file_out = ctx.obj['file_out']    
    outPath = ctx.obj['outPath']

    # load file_in and select sheet
    book = openpyxl.load_workbook(file_in)
    sheet = book.active

    click.echo('Adding ticker %s and saving to %s...' % (ticker, file_out))

    # check if already on sheet
    if findTicker(sheet, ticker)==False:

        # find first available row
        row = availableRow(sheet)

        # get ticker info
        price, div = priceDiv(ticker)

        # place data
        placeData(sheet, row, ticker, price, div)

        #update date
        updateDate(sheet)

        # save as file_out and close file_in
        book.save(file_out)
        book.close()

        click.echo('Ticker %s added!' % ticker)
        os.startfile(outPath)

    else:

        click.echo('Ticker %s is already in %s' % (ticker, file_out))
        
        # save as file_out and close file_in
        book.save(file_out)
        book.close()
    
# add_list adds a list of new tickers and updates just that ticker
@cli.command()
@click.argument('tickers', nargs=-1)
@click.pass_context
def addList(ctx, tickers):
    """Takes a list of tickers and adds them to file_in, saving to file_out"""

    file_in = ctx.obj['file_in']
    file_out = ctx.obj['file_out']    
    outPath = ctx.obj['outPath']

    # load file_in and select sheet
    book = openpyxl.load_workbook(file_in)
    sheet = book.active

    click.echo('Adding list of tickers...')

    # loop through all tickers in ticker
    if len(tickers) != 0:
        for ticker in tickers:

            # check if already on sheet
            if findTicker(sheet, ticker)==False:

                # find first available row
                row = availableRow(sheet)

                # get ticker info
                price, div = priceDiv(ticker)

                # place data
                placeData(sheet, row, ticker, price, div)

                click.echo('Ticker %s added!' % ticker)

                book.save(file_out)

            else:
                click.echo('Ticker %s already in %s.' % (ticker, file_out))

    # update date
    updateDate(sheet)

    click.echo('All tickers added.')

    # save to out_file and close
    book.save(file_out)
    book.close()

    os.startfile(outPath)

# delete command deletes a ticker from a sheet if the sheet exists
# and the ticker is present
@cli.command()
@click.argument('ticker')
@click.pass_context
def delete(ctx, ticker):
    """Deletes an extant ticker from file_in and saves to file_out"""

    file_in = ctx.obj['file_in']
    file_out = ctx.obj['file_out']
    outPath = ctx.obj['outPath']
    
    click.echo('Deleting %s and saving to %s...' % (ticker, file_out))

    # load file in and select sheet
    book = openpyxl.load_workbook(file_in)
    sheet = book.active

    # find ticker
    row = findTicker(sheet, ticker)

    # if ticker is not found
    if row == False:

        click.echo('Ticker %s not found...' % ticker)

    else:
        
        rowID = row[0].row
        sheet.delete_rows(rowID)
        click.echo('%s deleted...' % ticker)
        updateDate(sheet)
        book.save(file_out)

    # close book
    book.close()
    os.startfile(outPath)

# delete_list takes a list of tickers and deletes them from a sheet
# if the sheet exists and the ticker is present
@cli.command()
@click.argument('tickers', nargs=-1)
@click.pass_context
def deleteList(ctx, tickers):
    """Takes a list of tickers and deletes examples from file_in and saves to file_out"""

    file_in = ctx.obj['file_in']
    file_out = ctx.obj['file_out']    
    outPath = ctx.obj['outPath']
    
    click.echo('Deleting list of tickers...')

    # load file in and select sheet
    book = openpyxl.load_workbook(file_in)
    sheet = book.active

    # loop through all tickers in ticker
    if len(tickers) != 0:
        for ticker in tickers:

            # find ticker
            row = findTicker(sheet, ticker)

            # if ticker is not found
            if row == False:

                click.echo('Ticker %s not found...' % ticker)

            else:
                rowID = row[0].row
                sheet.delete_rows(rowID)
                click.echo('%s deleted...' % ticker)

    else:
        click.echo('delete_list requires a list containing at least one ticker')

    # save to file_out and close book
    updateDate(sheet)
    book.save(file_out)
    book.close()
    os.startfile(outPath)

# non-cli methods

# dividendUpdate takes two file names as strings
# looks fileIn and iterates through each row
# calling priceDiv if there is a ticker present
# Places information in correct position and saves to fileOut
def dividendUpdate(fileIn, fileOut):

    # open workbook at fileIn and select active sheet
    book = openpyxl.load_workbook(fileIn)
    sheet = book.active

    # count for tracking updates
    count = 0

    # iterate through all rows that contain tickers
    for row in sheet:

        # check to see if header or ticker
        if row[0].value != "Stock ID":

            # set ticker to ticker
            ticker = row[0].value

            # query yahoo_fin
            price, div = priceDiv(ticker)

            if price is not None:
                # send data to place data
                placeData(sheet, row, ticker, price, div)

            # update count
            count += 1

    # place update date
    updateDate(sheet)

    # print count after completed
    click.echo('Updated %d tickers!' % count)

    # save workbook at fileOut and close fileIn
    book.save(fileOut)
    book.close()

# priceDiv takes a ticker and queries yahoo_fin
# returns a tuple of price and dividend yield
def priceDiv(ticker):
    try:
        price = si.get_live_price(ticker)
        div = si.get_quote_table(ticker)['Forward Dividend & Yield'][:4]
        return price, div
    except:
        raise Exception('There was a problem updating %s...' % (ticker))

# places data in appropriate cells 
def placeData(sheet, row, ticker, price, div):

    # aquire the row id number
    rowID = row[0].row

    # generate strings for cols C, E, and F for rowID
    cRow = 'C%s' % rowID
    eRow = 'E%s' % rowID
    fRow = 'F%s' % rowID

    # generate strings for functions at C[rowID] and E[rowID]
    cFun = '=D%s/B%s' % (rowID, rowID)
    eFun = '=B%s/D%s' % (rowID, rowID)

    # place strings for functions at C[rowID] and E[rowID]
    sheet[cRow].value = cFun
    sheet[eRow].value = eFun    
    
    # store ticker in ticker
    row[0].value = ticker

    # update spreadhseet with new price and yield info
    row[1].value = price
    row[3].value = div

    # calculate shares per $1000
    shares = math.floor(1000/price)

    # generate function to calculate annual div yield per $1000 of shares
    fFun = '=%f * %f' % (float(shares), float(div))

    # places function
    sheet[fRow].value = fFun

# takes a sheet formatted for dividend tracking and updates the date
def updateDate(sheet):
    
    # set update date
    # collect update date
    now = datetime.datetime.now()
    nowShort = str(now)[:9]

    # place now string in H1
    sheet['H1'].value = str(now)

# returns first empty row
def availableRow(sheet):

    # count of rows starts at 1
    # used to count the number of occupied rows
    count = 1

    # iterates through each occupied row in the sheet
    for row in sheet:

        # checks to see if the ticker value for a row is None
        if row[0].value != None:

            # if there is a ticker, increase the count by 1
            count += 1

    # return sheet[count + 1] 
    return sheet[count]
        
# returns first row containing ticker
def findTicker(sheet, ticker):

    # checks each populated row of the sheet
    for row in sheet:

        # checks to see if the first entry in row matches ticker
        if row[0].value == ticker:

            # returns row if a match is found
            return row

    # returns False if no match is found
    return False

if __name__=='__main__':
    cli(obj={})
    
